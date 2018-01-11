from apiclient import errors
from django.utils.encoding import smart_str
import email
import base64
from htmlparser import strip_tags
from Sentiment import calEmotionalLevel
from openpyxl import load_workbook
import pandas as pd
import re
from util import construct_thread_url
from custom_classfication_module import CustomClassification
from DBHandler import get_not_downloaded_thread_ids, update_thread_status

clf = CustomClassification(False)

def get_mail_messageID(mail):
    return  mail['Message-Id']

def get_mail_subject(mail):
    return  mail['subject']

def get_mail_sender(mail):
    return  mail['From']

def get_mail_date(mail):
    return  mail['date']

def get_mail_receiver(mail):
    return  mail['To']

def get_mail_body(mail):
    data = ""
    if mail.is_multipart():
        for payload in mail.get_payload():
            # if payload.is_multipart(): ...
            data = data + str(payload)

    else:
        data = mail.get_payload()
    return data

def get_mail_text_body(mail):
    return mail.get_payload()[0]


def get_thread(gmail,thread_id,fmat):
    """Get a Thread.

     Args:
       thread_id: The ID of the Thread required.

     Returns:
       Thread with matching ID.
     """

    try:
        thread = gmail.users().threads().get(userId='me', id=thread_id,format=fmat).execute()
        messages = thread[smart_str('messages')]
        #print('thread id: '+thread[smart_str('id')]+'- number of messages ''in this thread:'+str(len(messages)))
        return thread
    except errors.HttpError, error:
        print
        'An error occurred: %s' % error

def get_mime_mail(service,mail_id):
    """Get a Message and use it to create a MIME Message.

    Args:
      service: Authorized Gmail API service instance.
      user_id: User's email address. The special value "me"
      can be used to indicate the authenticated user.
      msg_id: The ID of the Message required.

    Returns:
      A MIME Message, consisting of data from Message.
    """
    try:
        message = service.users().messages().get(userId='me', id=mail_id,
                                                 format='raw').execute()

        msg_str = base64.urlsafe_b64decode(message['raw'].encode('ASCII'))

        mime_msg = email.message_from_string(msg_str)

        return mime_msg
    except errors.HttpError, error:
        print
        'An error occurred: %s' % error

def ListThreadsMatchingQuery(service, user_id, query=''):
    """List all Threads of the user's mailbox matching the query.

    Args:
      service: Authorized Gmail API service instance.
      user_id: User's email address. The special value "me"
      can be used to indicate the authenticated user.
      query: String used to filter messages returned.
             Eg.- 'label:UNREAD' for unread messages only.

    Returns:
      List of threads that match the criteria of the query. Note that the returned
      list contains Thread IDs, you must use get with the appropriate
      ID to get the details for a Thread.
    """
    try:
        response = service.users().threads().list(userId=user_id, q=query).execute()
        threads = []
        if 'threads' in response:
            threads.extend(response['threads'])

        while 'nextPageToken' in response:
            page_token = response['nextPageToken']
            response = service.users().threads().list(userId=user_id, q=query,
                                                      pageToken=page_token).execute()
            threads.extend(response['threads'])

        return threads
    except errors.HttpError, error:
        print 'An error occurred: %s' % error


def ListThreadsWithLabels(service, user_id, label_ids=[]):
    """List all Threads of the user's mailbox with label_ids applied.

    Args:
      service: Authorized Gmail API service instance.
      user_id: User's email address. The special value "me"
      can be used to indicate the authenticated user.
      label_ids: Only return Threads with these labelIds applied.

    Returns:
      List of threads that match the criteria of the query. Note that the returned
      list contains Thread IDs, you must use get with the appropriate
      ID to get the details for a Thread.
    """

    try:
        response = service.users().threads().list(userId=user_id,labelIds=label_ids,q='after:2017/12/01 before:2018/01/01').execute()
        threads = []
        if 'threads' in response:
            threads.extend(response['threads'])

        while 'nextPageToken' in response:
            page_token = response['nextPageToken']
            response = service.users().threads().list(userId=user_id,
                                                      labelIds=label_ids,
                                                      pageToken=page_token).execute()
            threads.extend(response['threads'])

        return threads
    except errors.HttpError, error:
        print 'An error occurred: %s' % error

def get_mail_body_word_count(body):
    return len(body.split(" "))

def read_mails_in_thread(gmail,thread):
    '''
    get theread and fettch email for each email id
    Args:
        thread: thread with emails ids

    Returns: list with data in emails


    cnx = mysql.connector.connect(user='root',password='houses123',host='127.0.0.1', database='emailing')
    cursor = cnx.cursor()
    '''

    # Loading results.xlsx sheet to store the results
    wb=load_workbook("/home/ching/WORK/SentimentAnalysis/results.xlsx")
    ws = wb.active

    processed_thread = {}
    processed_thread["thread_id"] = thread
    processed_thread["message_count"] = len(thread[smart_str('messages')])

    numofMsgs = processed_thread["message_count"]

    #print("num of msgs:--   "+ str(numofMsgs))
    #print ("threadid -- "+ str(thread[smart_str('id')]))

    emotionalList = []
    infoextract = []

    subjectCol = 'A'
    linkofMessageCol = 'B'
    dateCol = 'C'
    senderCol = 'D'
    complexityCol = 'E'
    sentimentCol = 'F'

    count = 0

    # retrieve each mail in raw format to read
    for mail in thread[smart_str('messages')]:
        emotionalList = []
        mail_id = smart_str('id')
        if mail_id is not None:
            raw_mail = get_mime_mail(gmail,mail[smart_str('id')])
            subject = get_mail_subject(raw_mail)
            date = get_mail_date(raw_mail)
            sender = get_mail_sender(raw_mail)
            mid = get_mail_messageID(raw_mail)
            msglink = construct_thread_url(mid)

            body = get_mail_body(raw_mail)
            processed_body = process_mail_body(body)
            processed_body = processed_body.split("You received this message because you are subscribed to the Google")[0]

            numofwords = len(processed_body.split())

            emotionalLevel = calEmotionalLevel(processed_body)

            complexity_data = pd.DataFrame([[numofMsgs, subject, processed_body, numofwords]],columns=['NO_OF_EMAIL_IN_THREAD', 'SUBJECT', 'EMAIL_BODY', 'WORD_COUNT']);
            complexity_level = int(clf.predict_instance(complexity_data)+1)


            ws[subjectCol+str(1+count)] = str(subject)
            ws[linkofMessageCol+str(1+count)] = str(msglink)
            ws[dateCol+str(1+count)] = str(date[4:16])
            ws[senderCol+str(1+count)] = str(sender)
            ws[complexityCol+str(1+count)] = str(complexity_level)
            ws[sentimentCol+str(1+count)] = str(emotionalLevel)
            # update DB state
            update_thread_status(thread)
            count = count + 1
        else:
            print("Msg id is None!. Current Thread id: " + thread['id'])
    wb.save("/home/ching/WORK/SentimentAnalysis/results.xlsx")

    return emotionalList



def get_mail_threads(gmail,threadId_list):
    '''
    get the content of mail threads when a thread id is given
    Args:
        threadId_lisget_mail_threadst: thread id

    Returns: list of mail threads with messages

    '''


    threads = []
    count = 0
    #iterate though each thread
    for thread in threadId_list:
        count = count + 1
        thread_data = get_thread(gmail,smart_str(thread),'minimal')
        processed_thread = read_mails_in_thread(gmail,thread_data)

        threads.append(processed_thread)

    return processed_thread

def process_mail_body(content):
    imageID = 0
    idIndex = ""
    nohtml = strip_tags(content)

    content = nohtml.split("From ")
    if(len(content))>1:
        content = content[1]
    else:
        content = content[0]
    content = ">".join(content.split(">>"))
    content = re.sub('>.*?\n','',content, flags=re.DOTALL)
    content = re.sub('Content-Transfer.*?\n','',content, flags=re.DOTALL)
    content = re.sub('obody.*? ','',content, flags=re.DOTALL)
    content = re.sub('>.*?\n','',content, flags=re.DOTALL)

    imageID = content.find("boundary=")
    idIndex = content[imageID + 10:imageID+38]

    new = content.replace(str(idIndex),'', 2)
    content = new[:new.find(str(idIndex))]
    content = re.sub('Content-Type:.*?\n','',content, flags=re.DOTALL)
    content = content.replace("--",'')

    body = ""
    length = len(content)
    for i in range(1,length):
        body = body+content[i]
    return body

def find_between(s):
    first = "u'id': u'"
    last = "', u"
    try:
        start = s.index( first ) + len( first )
        end = s.index( last, start )
        return s[start:end]
    except ValueError:
        return ""



