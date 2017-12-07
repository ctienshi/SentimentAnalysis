from __future__ import division
from pandas_ml import ConfusionMatrix
import matplotlib.pyplot as plt
from openpyxl import load_workbook

wb=load_workbook("/home/ching/Downloads/dataset.xlsx")
ws = wb.active
actual = ws['A']
pred = ws['B']
act_arr = []
pred_arr = []
print  (actual[3].value)
for i in range(100):
    act_arr.append(actual[i].value)
    pred_arr.append(pred[i].value)
#y_true = [2, 1, 0, 0, -2, -1, 2, 2, 1, -2, -2, -1]
#y_pred = [1, 1, 0, 1, -2, 1, 2, 2, 1, 2, 2, 1]
confusion_matrix = ConfusionMatrix(act_arr, pred_arr)
print("Confusion matrix:\n%s" % confusion_matrix)
confusion_matrix.plot()
plt.show()

def perf_measure(y_actual, y_hat):
    cor = 0
    wro = 0

    for i in range(len(y_hat)):
        if y_actual[i]==y_hat[i]:
            cor += 1
        else:
            wro += 1
    a = cor/(cor+wro)
    return(a)

print ("\n")
a = perf_measure(act_arr,pred_arr)
print("The Accuray is: "+str(a*100)+"%")