from apiclient import errors
from django.utils.encoding import smart_str
import email
import base64
from htmlparser import strip_tags
from Sentiment import calEmotionalLevel
from openpyxl import load_workbook
import mysql.connector
import re

def get_mail_subject(mail):
    return  mail['subject']

def get_mail_sender(mail):
    return  mail['From']

def get_mail_date(mail):
    return  mail['date']

def get_mail_receiver(mail):
    return  mail['To']

def get_mail_body(mail):
    data = ""
    if mail.is_multipart():
        for payload in mail.get_payload():
            # if payload.is_multipart(): ...
            data = data + str(payload)

    else:
        data = mail.get_payload()
    return data

def get_mail_text_body(mail):
    return mail.get_payload()[0]


def get_thread(gmail,thread_id,fmat):
    """Get a Thread.

     Args:
       thread_id: The ID of the Thread required.

     Returns:
       Thread with matching ID.
     """

    try:
        thread = gmail.users().threads().get(userId='me', id=thread_id,format=fmat).execute()
        messages = thread[smart_str('messages')]
        print('thread id: '+thread[smart_str('id')]+'- number of messages '
                                                    'in this thread:'+str(len(messages)))
        return thread
    except errors.HttpError, error:
        print
        'An error occurred: %s' % error

def get_mime_mail(service,mail_id):
    """Get a Message and use it to create a MIME Message.

    Args:
      service: Authorized Gmail API service instance.
      user_id: User's email address. The special value "me"
      can be used to indicate the authenticated user.
      msg_id: The ID of the Message required.

    Returns:
      A MIME Message, consisting of data from Message.
    """
    try:
        message = service.users().messages().get(userId='me', id=mail_id,
                                                 format='raw').execute()

        msg_str = base64.urlsafe_b64decode(message['raw'].encode('ASCII'))

        mime_msg = email.message_from_string(msg_str)

        return mime_msg
    except errors.HttpError, error:
        print
        'An error occurred: %s' % error

def ListThreadsMatchingQuery(service, user_id, query=''):
    """List all Threads of the user's mailbox matching the query.

    Args:
      service: Authorized Gmail API service instance.
      user_id: User's email address. The special value "me"
      can be used to indicate the authenticated user.
      query: String used to filter messages returned.
             Eg.- 'label:UNREAD' for unread messages only.

    Returns:
      List of threads that match the criteria of the query. Note that the returned
      list contains Thread IDs, you must use get with the appropriate
      ID to get the details for a Thread.
    """
    try:
        response = service.users().threads().list(userId=user_id, q=query).execute()
        threads = []
        if 'threads' in response:
            threads.extend(response['threads'])

        while 'nextPageToken' in response:
            page_token = response['nextPageToken']
            response = service.users().threads().list(userId=user_id, q=query,
                                                      pageToken=page_token).execute()
            threads.extend(response['threads'])

        return threads
    except errors.HttpError, error:
        print 'An error occurred: %s' % error


def ListThreadsWithLabels(service, user_id, label_ids=[]):
    """List all Threads of the user's mailbox with label_ids applied.

    Args:
      service: Authorized Gmail API service instance.
      user_id: User's email address. The special value "me"
      can be used to indicate the authenticated user.
      label_ids: Only return Threads with these labelIds applied.

    Returns:
      List of threads that match the criteria of the query. Note that the returned
      list contains Thread IDs, you must use get with the appropriate
      ID to get the details for a Thread.
    """

    try:
        response = service.users().threads().list(userId=user_id,labelIds=label_ids,q='after:2017/12/01 before:2018/01/01').execute()
        threads = []
        if 'threads' in response:
            threads.extend(response['threads'])

        while 'nextPageToken' in response:
            page_token = response['nextPageToken']
            response = service.users().threads().list(userId=user_id,
                                                      labelIds=label_ids,
                                                      pageToken=page_token).execute()
            threads.extend(response['threads'])

        return threads
    except errors.HttpError, error:
        print 'An error occurred: %s' % error

def get_mail_body_word_count(body):
    return len(body.split(" "))

def read_mails_in_thread(gmail,thread):
    '''
    get theread and fettch email for each email id
    Args:
        thread: thread with emails ids

    Returns: list with data in emails

    '''
    cnx = mysql.connector.connect(user='root',password='houses123',host='127.0.0.1', database='testing')
    cursor = cnx.cursor()

    processed_thread = {}
    processed_thread["thread_id"] = thread
    processed_thread["message_count"] = len(thread[smart_str('messages')])
    data = []
    emotionalList = []
    lastmessage = ""
    count = 0
    vpos = 0
    pos = 0
    neu = 0
    neg = 0
    vneg = 0
    overall = ""
    # retrie each mail in raw format for read
    for mail in thread[smart_str('messages')]:
        emotionalList = []

        mail_data = {}
        raw_mail = get_mime_mail(gmail,mail[smart_str('id')])
        subject = get_mail_subject(raw_mail)
        date = get_mail_date(raw_mail)
        sender = get_mail_sender(raw_mail)
        if count == 0:
            receiver = get_mail_receiver(raw_mail)
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++-----------------------------")
            print (receiver)
        count = count + 1
        body = get_mail_body(raw_mail)
        processed_body = process_mail_body(body)
        processed_body = processed_body.split("You received this message because you are subscribed to the Google")[0]

        emotionalLevel = calEmotionalLevel(processed_body)

        if emotionalLevel == 2:
            vpos = vpos + 1
        elif emotionalLevel == 1:
            pos = pos + 1
        elif emotionalLevel == 0:
            neu = neu + 1
        elif emotionalLevel == -1:
            neg = neg + 1
        elif emotionalLevel == -2:
            vneg = vneg + 1
        mail_data["id"] = mail
        mail_data["subject"] = subject
        mail_data["body"] = processed_body
        data.append(mail_data)

        lastmessage = processed_body

    if vneg > 0:
        overall = "Very Negative"

    elif neg > 0:
        overall = "Negative"

    elif vpos > 0:
        overall = "Very Positive"

    elif pos > 0:
        overall = "Positive"

    else:
        overall = "Neutral"
    processed_thread["mail-list"] = data
    emotionalList.append(date)
    emotionalList.append(subject)
    emotionalList.append(vpos)
    emotionalList.append(pos)
    emotionalList.append(neu)
    emotionalList.append(neg)
    emotionalList.append(vneg)
    emotionalList.append(overall)
    emotionalList.append(sender)
    emotionalList.append(lastmessage)
    emotionalList.append(emotionalLevel)
    emotionalList.append(receiver)
    print("+++++++++++++++++++++++++++++++++++LAST++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    #print (lastmessage[:24])

    print (emotionalList)
    #print (emotionalList[6])
    '''
    add_employee = ("INSERT INTO emails "
                    "(threadid,date, sub, vpositive, positive, neutral,negative,vnegative,overall,sender,lastmsg,emotionalLevel,group) "
                    "VALUES (%s, %s, %s, %s, %s,%s,%s)")
    data_employee = (thread[smart_str('id')], emotionalList[0], emotionalList[1],emotionalList[2],emotionalList[3],emotionalList[4],emotionalList[5])
    

    cursor.execute(add_employee, data_employee)
    cnx.commit()

    cursor.close()
    cnx.close()
    '''
    return emotionalList

def get_mail_threads(gmail,threadId_list):
    '''
    get the content of mail threads when a thread id is given
    Args:
        threadId_lisget_mail_threadst: thread id

    Returns: list of mail threads with messages

    '''
    wb=load_workbook("/home/ching/WORK/SentimentAnalysis/bizdev.xlsx")
    ws = wb.active

    threads = []
    count = 0
    #iterate though each thread
    for thread in threadId_list:
        thread_data = get_thread(gmail,smart_str(thread),'minimal')
        processed_thread = read_mails_in_thread(gmail,thread_data)
        subjectTitle = 'A'
        vpos = 'B'
        pos = 'C'
        neu = 'D'
        neg = 'E'
        vneg = 'F'
        threads.append(processed_thread)
        ws[subjectTitle+str(1+count)] = str(processed_thread[0])
        ws[vpos+str(1+count)] = str(processed_thread[1])
        ws[pos+str(1+count)] = str(processed_thread[2])
        ws[neu+str(1+count)] = str(processed_thread[3])
        ws[neg+str(1+count)] = str(processed_thread[4])
        ws[vneg+str(1+count)] = str(processed_thread[5])
        count = count + 1
        #if count == len(threadId_list):
        if count == 1:
            break

    wb.save("/home/ching/WORK/SentimentAnalysis/bizdev.xlsx")
    return processed_thread

def process_mail_body(content):
    imageID = 0
    idIndex = ""
    nohtml = strip_tags(content)

    content = nohtml.split("From ")
    if(len(content))>1:
        content = content[1]
    else:
        content = content[0]
    content = ">".join(content.split(">>"))
    content = re.sub('>.*?\n','',content, flags=re.DOTALL)
    content = re.sub('Content-Transfer.*?\n','',content, flags=re.DOTALL)
    content = re.sub('obody.*? ','',content, flags=re.DOTALL)
    #content = re.sub('<div.*? </div>','',content, flags=re.DOTALL)
    content = re.sub('>.*?\n','',content, flags=re.DOTALL)
    #content = content.split("<div")[0].split("\n\n")

    imageID = content.find("boundary=")
    idIndex = content[imageID + 10:imageID+38]
    #print ("==========================================================ID============================================================")
    #print (idIndex)
    new = content.replace(str(idIndex),'', 2)
    content = new[:new.find(str(idIndex))]
    content = re.sub('Content-Type:.*?\n','',content, flags=re.DOTALL)
    content = content.replace("--",'')

    #data = content.split(">")[0].split("\n\n")

    body = ""
    length = len(content)
    for i in range(1,length):
        body = body+content[i]
    return body

def find_between(s):
    first = "u'id': u'"
    last = "', u"
    try:
        start = s.index( first ) + len( first )
        end = s.index( last, start )
        return s[start:end]
    except ValueError:
        return ""



