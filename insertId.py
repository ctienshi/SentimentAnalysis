from apiclient import discovery
from httplib2 import Http
from oauth2client import file, client, tools
from extractImportantv1 import ListThreadsMatchingQuery
from extractImportantv1 import get_mail_threads
from openpyxl import load_workbook
from DBHandler import put_thread_id,get_not_downloaded_thread_ids
SCOPES = 'https://www.googleapis.com/auth/gmail.modify' # we are using modify and not readonly, as we will be marking the messages Read
store = file.Storage('storage.json')
creds = store.get()

if not creds or creds.invalid:
    flow = client.flow_from_clientsecrets('client_secret.json', SCOPES)
    creds = tools.run_flow(flow, store)
GMAIL = discovery.build('gmail', 'v1', http=creds.authorize(Http()))

user_id =  'me'

#Getting the threadID list for a particular label
threadids = []
wb=load_workbook("/home/ching/WORK/SentimentAnalysis/data.xlsx")
ws = wb.active

# Query to extract the list of threadids for the last 3 months
ListofThreads = ListThreadsMatchingQuery(GMAIL,'me','after:2017/10/01 before:2017/12/31')

ColumnId = 'F'

arrayofThreads = []


#To save the extracted threadId list to an excel sheet (The column 'F')
'''
for i in range(len(ListofThreads)):
    print (count)
    threadid = find_between(str(ListofThreads[i]))
    count = count + 1
    ws[ColumnId+str(1+count)] = threadid
'''


for i in range(len(ListofThreads)):
    print (count)
    threadid = find_between(str(ListofThreads[i]))
    count = count + 1
    # save to DB. Since it's not downloaded yet set Downloaded=False
    put_thread_id(threadid, False)
    # ws[ColumnId+str(1+count)] = threadid

# Extracts the threadids from the excel sheet and loads to a list (1000 threadids)
limit = 1000
for count in range(limit):
    #column = ws[ColumnId+str(1+count)]
    data = get_not_downloaded_thread_ids(limit)
    threadid = data['_id']
    # print(column.value)
    arrayofThreads.append(str(threadid))


# Get the corresponding data for the thread list
get_mail_threads(GMAIL,arrayofThreads)


wb.save("/home/ching/WORK/SentimentAnalysis/data.xlsx")



