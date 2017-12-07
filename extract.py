#java -mx4g -cp "*" edu.stanford.nlp.pipeline.StanfordCoreNLPServer -port 9000 -timeout 15000
from Sentiment import calEmotionalLevel
from Sentiment import tes
from openpyxl import load_workbook
wb=load_workbook("/home/ching/WORK/SentimentAnalysis/testData/dataset.xlsx")

ws = wb.active
first_column = ws['E']
col = ws['B']
test = str(first_column[4].value)

word = "You received this message because"

for i in range(90,len(first_column)):
    a = str(first_column[i].value)
    a = a[:a.find(word)]

    x = calEmotionalLevel(a)
    p = tes(a)
    column_cell = 'B'
    cell = 'C'
    ws[column_cell+str(i+1)] = x
    ws[cell+str(i+1)] = p

    print(str(i+1)+" " +str(x)) + " " + str(p)
    print ("\n")


wb.save("/home/ching/WORK/SentimentAnalysis/testData/dataset.xlsx")

